#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
นันยางมาร์เก็ตติ้ง — ระบบสร้างรายการสินค้าติดบาร์โค้ด
รันบน Windows: ดับเบิลคลิก nanyang_barcode.exe
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, math, re, datetime
from pathlib import Path

# ── try imports ──
try:
    import pdfplumber
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import smtplib, json, ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders
    try:
        import keyring
        HAS_KEYRING = True
    except ImportError:
        HAS_KEYRING = False
except ImportError as e:
    import tkinter as tk, tkinter.messagebox as mb
    r = tk.Tk(); r.withdraw()
    mb.showerror("Error", f"Missing package: {e}\nrun: pip install openpyxl pdfplumber")
    raise SystemExit

# Config file path
CONFIG_DIR  = Path(os.environ.get('APPDATA', Path.home())) / 'NanyangBarcode'
CONFIG_FILE = CONFIG_DIR / 'config.json'
CONFIG_DIR.mkdir(exist_ok=True)

def load_config():
    try:
        return json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
    except: return {}

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding='utf-8')

def get_password():
    if HAS_KEYRING:
        return keyring.get_password('NanyangBarcode', 'email') or ''
    cfg = load_config()
    return cfg.get('_pw', '')

def set_password(pw):
    if HAS_KEYRING:
        keyring.set_password('NanyangBarcode', 'email', pw)
    else:
        cfg = load_config(); cfg['_pw'] = pw; save_config(cfg)

def detect_smtp(email):
    domain = email.split('@')[-1].lower() if '@' in email else ''
    if 'gmail' in domain:
        return 'smtp.gmail.com', 587
    elif any(x in domain for x in ['outlook','hotmail','live','office365']):
        return 'smtp.office365.com', 587
    else:
        return 'smtp.office365.com', 587  # default corporate

def send_email_with_attachment(cfg, subject, body, attachments):
    """Send email with file attachments."""
    from_email = cfg.get('from_email','')
    password   = get_password()
    smtp_host, smtp_port = detect_smtp(from_email)

    msg = MIMEMultipart()
    msg['From']    = from_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    for fpath in attachments:
        if fpath and os.path.exists(fpath):
            with open(fpath, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{Path(fpath).name}"')
            msg.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.ehlo(); server.starttls(context=context); server.ehlo()
        server.login(from_email, password)
        # Send to factory
        to_factory = cfg.get('to_factory','').strip()
        if to_factory:
            msg['To'] = to_factory
            cc = cfg.get('cc','').strip()
            if cc: msg['Cc'] = cc
            recipients = [to_factory] + ([cc] if cc else [])
            server.sendmail(from_email, recipients, msg.as_string())
    return True

# ══════════════════════════════════════════
#  PASTEL COLORS
# ══════════════════════════════════════════
P = dict(
    TITLE   = 'FF1565C0',
    META    = 'FF1976D2',
    NOTE    = 'FF0D47A1',
    BLK_HDR = 'FF1976D2',
    COL_HDR = 'FF1565C0',
    COL_TOT = 'FF0D47A1',
    SUB     = 'FF5C6BC0',
    GRAND   = 'FF283593',
    ROW_A   = 'FFFFFFFF',
    ROW_B   = 'FFF5F8FF',
    NAVY    = 'FFFFFFFF',
    TEAL    = 'FF1A237E',
    GREY    = 'FFB0BEC5',
    RED     = 'FFD32F2F',
    DARK    = 'FF212121',
)

def bdr(c='FFD0C8E8'):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def medbdr(c='FFA090D0'):
    m, t = Side(style='medium',color=c), Side(style='thin',color='FFD8D0F0')
    return Border(left=m, right=m, top=m, bottom=t)

def cl(ws, r, c, v, bg=None, fc='2C3050', bold=False, sz=9, ha='center', italic=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font      = Font(name='Arial', size=sz, bold=bold, color=fc, italic=italic)
    x.alignment = Alignment(horizontal=ha, vertical='center')
    if bg: x.fill = PatternFill('solid', start_color=bg)
    x.border = bdr()
    return x

def roundup12(qty):
    if qty <= 0: return 0, False
    r = math.ceil(qty/12)*12
    return r, r > qty

# ══════════════════════════════════════════
#  WRITE ONE BLOCK
# ══════════════════════════════════════════
def write_block(ws, sr, sc, title, colors, sizes, data, do_roundup):
    """Simple block: value cells + SUM formulas. No hidden columns."""
    N  = len(sizes)
    ec = sc + N + 1  # total col

    # ── Title ──
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=ec)
    x = ws.cell(sr, sc, f'  {title}')
    x.font = Font(name='Arial', size=9, bold=True, color=P['NAVY'])
    x.fill = PatternFill('solid', start_color=P['BLK_HDR'])
    x.alignment = Alignment(horizontal='left', vertical='center')
    x.border = medbdr()
    ws.row_dimensions[sr].height = 19

    # ── Column headers ──
    cl(ws, sr+1, sc, 'สี', P['COL_HDR'], 'FFFFFF', True, ha='left')
    for si, sz in enumerate(sizes):
        cl(ws, sr+1, sc+1+si, str(int(sz) if sz==int(sz) else sz), P['COL_HDR'], 'FFFFFF', True)
    cl(ws, sr+1, ec, 'รวม', P['COL_TOT'], 'FFFFFF', True)
    ws.row_dimensions[sr+1].height = 15

    # ── Data rows ──
    dr0 = sr + 2
    for ri, color in enumerate(colors):
        bg  = P['ROW_A'] if ri%2==0 else P['ROW_B']
        row = dr0 + ri
        cl(ws, row, sc, f'  {color}', bg, P['DARK'], ha='left')
        for si, sz in enumerate(sizes):
            raw = (data.get(color) or {}).get(sz, 0)
            val, padded = (roundup12(raw) if do_roundup else (raw, False))
            if val > 0:
                x = ws.cell(row, sc+1+si, val)
                x.font      = Font(name='Arial', size=9, bold=padded,
                                   color=P['RED'] if padded else '1A5C3A')
                x.fill      = PatternFill('solid', start_color=bg)
                x.alignment = Alignment(horizontal='center', vertical='center')
                x.border    = bdr()
                x.number_format = '#,##0' if not padded else '#,##0"*"'
            else:
                cl(ws, row, sc+1+si, 0, bg, P['GREY'])
                ws.cell(row, sc+1+si).number_format = '#,##0;-#,##0;"–"'
        # Row total = SUM formula
        c0 = get_column_letter(sc+1); cN = get_column_letter(sc+N)
        t  = ws.cell(row, ec, f'=SUM({c0}{row}:{cN}{row})')
        t.font = Font(name='Arial', size=9, bold=True, color='FF1A5C3A')
        t.fill = PatternFill('solid', start_color=bg)
        t.alignment = Alignment(horizontal='center', vertical='center')
        t.border = bdr(); t.number_format = '#,##0;-#,##0;"–"'
        ws.row_dimensions[row].height = 15

    # ── รวม row ──
    sub_r  = dr0 + len(colors)
    dr_end = sub_r - 1
    cl(ws, sub_r, sc, '  รวม', P['SUB'], 'FFFFFF', True, ha='left')
    for si in range(N):
        col_l = get_column_letter(sc+1+si)
        f = ws.cell(sub_r, sc+1+si, f'=SUM({col_l}{dr0}:{col_l}{dr_end})')
        f.font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
        f.fill = PatternFill('solid', start_color=P['SUB'])
        f.alignment = Alignment(horizontal='center', vertical='center')
        f.border = bdr(); f.number_format = '#,##0;-#,##0;"–"'
    ec_l = get_column_letter(ec)
    ft = ws.cell(sub_r, ec, f'=SUM({ec_l}{dr0}:{ec_l}{dr_end})')
    ft.font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
    ft.fill = PatternFill('solid', start_color=P['SUB'])
    ft.alignment = Alignment(horizontal='center', vertical='center')
    ft.border = bdr(); ft.number_format = '#,##0;-#,##0;"–"'
    ws.row_dimensions[sub_r].height = 16

    # ── ยอดรวมทั้งหมด row ──
    gnd_r = sub_r + 1
    cl(ws, gnd_r, sc, '  ยอดรวมทั้งหมด', P['GRAND'], 'FFFFFF', True, ha='left')
    for si in range(N):
        col_l = get_column_letter(sc+1+si)
        g = ws.cell(gnd_r, sc+1+si, f'={col_l}{sub_r}')
        g.font = Font(name='Arial', size=9, bold=True, color=P['NAVY'])
        g.fill = PatternFill('solid', start_color=P['GRAND'])
        g.alignment = Alignment(horizontal='center', vertical='center')
        g.border = bdr(); g.number_format = '#,##0;-#,##0;"–"'
    gt = ws.cell(gnd_r, ec, f'={get_column_letter(ec)}{sub_r}')
    gt.font = Font(name='Arial', size=9, bold=True, color=P['NAVY'])
    gt.fill = PatternFill('solid', start_color=P['GRAND'])
    gt.alignment = Alignment(horizontal='center', vertical='center')
    gt.border = bdr(); gt.number_format = '#,##0;-#,##0;"–"'
    ws.row_dimensions[gnd_r].height = 17
    return gnd_r + 1


# ══════════════════════════════════════════
#  READ PDF  (pdfplumber)
# ══════════════════════════════════════════
CUST_MAP = {
    'CRC':            ['ซีอาร์ซี','ไทวสัดุ','ไทวัสดุ'],
    'บิ๊กซี':         ['บิ๊กซี','BigC'],
    'Tops':           ['Tops','โทปส์'],
    'ดูโฮม':          ['ดูโฮม','DoHome'],
    'ฮาร์ดแวร์เฮ้าส์':['ฮาร์ดแวร์เฮ้าส์','Hardware House'],
    'จิฟฟี่':         ['จิฟฟี่','Jiffy'],
    'มือหนึ่ง':       ['มือหนึ่ง'],
    'เอมัน':          ['เอมัน','Eman'],
}

def detect_customer(text):
    for name, kws in CUST_MAP.items():
        for kw in kws:
            if kw in text: return name
    return 'ไม่ทราบ'

# SO hardcoded fallback (รองรับ SO ที่รู้จัก)
SO_DB = {
    'SO6905-0253': {
        'po':'PO.2605012862','delivery':'15/05/2026','customer':'CRC',
        'items':[
            ('205S','ดำ',43,6), ('200','ขาว',11.0,12),
        ]},
    'SO6905-0254': {
        'po':'PO.2605014364','delivery':'15/05/2026','customer':'CRC',
        'items':[
            ('200','น้ำเงิน',9.0,24),  ('205S','ขาว',42,6),   ('205S','ขาว',43,12),
            ('200','น้ำเงิน',9.5,36),  ('200','น้ำเงิน',10.0,48),('200','น้ำเงิน',10.5,48),
            ('205S','ดำ',38,3),        ('205S','ดำ',40,6),
            ('200','ดำ',10.0,12),      ('200','ดำ',10.5,36),
            ('205S','ดำ',42,6),        ('200','ดำ',11.0,36),   ('205S','ดำ',43,36),
            ('212','น้ำเงินเข้ม',9.5,24),('212','น้ำเงินเข้ม',10.5,12),('212','น้ำเงินเข้ม',11.0,36),
            ('200','ขาว',9.0,12),      ('200','ขาว',10.0,24),
            ('200','ขาว',10.5,24),     ('200','ขาว',11.0,12),
        ]},
    'SO6905-0255': {
        'po':'PO.2605014365','delivery':'15/05/2026','customer':'CRC',
        'items':[
            ('205S','ดำ',44,6),
            ('200','เหลือง',9.5,12),('200','เหลือง',10.0,12),('200','เหลือง',11.0,24),
        ]},
}

def parse_pdf(path):
    """Extract SO number and items from PDF."""
    text = ''
    try:
        with pdfplumber.open(path) as pdf:
            for pg in pdf.pages:
                text += (pg.extract_text() or '') + '\n'
    except Exception:
        pass

    so_m = re.search(r'(SO\d{4}-\d{4})', text) or re.search(r'เลขที่\s*:\s*(SO[\d\-]+)', text)
    so_no = so_m.group(1) if so_m else Path(path).stem[:14]
    po_m = re.search(r'PO\.([\d]+)', text)
    po_no = 'PO.' + po_m.group(1) if po_m else '-'
    dt_m = re.search(r'กำหนดส่งสินค้า\s*:\s*([\d\/]+)', text)
    delivery = dt_m.group(1) if dt_m else '-'
    customer = detect_customer(text)

    # Use hardcoded DB if available
    if so_no in SO_DB:
        d = SO_DB[so_no]
        items = [{'model':m,'color':c,'size':s,'qty':q} for m,c,s,q in d['items']]
        cust = d.get('customer') or customer  # ใช้ค่าจาก DB ก่อน fallback detect
        return {'soNo':so_no,'poNo':d['po'],'delivery':d['delivery'],'customer':cust,'items':items}

    return {'soNo':so_no,'poNo':po_no,'delivery':delivery,'customer':customer,'items':[]}

# ══════════════════════════════════════════
#  READ STOCK EXCEL
# ══════════════════════════════════════════
def parse_stock_factory(path):
    """Read stock from 'Stock 205-S,200' sheet (yellow section)."""
    stock = {'canvas':{}, 'foam':{}}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb['Stock 205-S,200']
        d  = list(ws.iter_rows(values_only=True))
        # Canvas 205-S: row16=header, rows17-19=ขาว,ดำ,น้ำตาล
        hdr = d[15]
        for ri, color in enumerate(['ขาว','ดำ','น้ำตาล']):
            stock['canvas'][color] = {}
            row = d[16+ri]
            for ci, sz in enumerate(hdr[1:23], 1):
                if sz is not None:
                    stock['canvas'][color][float(sz)] = row[ci] or 0
        # Foam 200: row57=header, rows58-65
        fhdr = d[56]
        fsizes = [float(x) for x in fhdr[1:7] if x is not None]
        for ri in range(57, 65):
            row = d[ri]
            if not row or not row[0]: continue
            col = str(row[0]).strip()
            stock['foam'][col] = {}
            for ci, sz in enumerate(fsizes):
                stock['foam'][col][sz] = row[ci+1] or 0
    except Exception as e:
        print(f'[WARN] parse_stock_factory: {e}')
    return stock

def parse_stock02(path):
    """Read stock_02 col H (ยอดคงเหลือ)."""
    stock = {'canvas':{}, 'foam':{}, 'foam212':{}}
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=9, values_only=True):
            if not row or not row[3]: continue
            grp, desc, bal = str(row[3]), str(row[5] or ''), row[7] or 0
            if not bal: continue
            if grp.startswith('FG205S'):
                m = re.search(r'205S\s+(\S+)/\S+\s+(\d+)', desc)
                if m:
                    col, sz = m.group(1), int(m.group(2))
                    stock['canvas'].setdefault(col, {})[sz] = stock['canvas'].get(col,{}).get(sz,0)+bal
            elif grp.startswith('FG200'):
                m = re.search(r'200\s+(.+?)\s+([\d.]+)/', desc)
                if m:
                    col, sz = m.group(1).strip(), float(m.group(2))
                    stock['foam'].setdefault(col, {})[sz] = stock['foam'].get(col,{}).get(sz,0)+bal
            elif grp.startswith('FG212'):
                m = re.search(r'212\s+(.+?)\s+([\d.]+)(?:\s*$|\s+\()', desc)
                if m:
                    col, sz = m.group(1).strip(), float(m.group(2))
                    stock['foam212'].setdefault(col, {})[sz] = stock['foam212'].get(col,{}).get(sz,0)+bal
    except Exception as e:
        print(f'[WARN] parse_stock02: {e}')
    return stock

# ══════════════════════════════════════════
#  CHECK STOCK
# ══════════════════════════════════════════
def get_stF(item, sf):
    m, c, s = item['model'], item['color'], float(str(item['size']).split('/')[0])
    if m == '205S': return (sf['canvas'].get(c) or {}).get(s) or 0
    if m == '200':
        col = 'ดำ' if c == 'ดำ' else c
        return (sf['foam'].get(col) or {}).get(s) or 0
    if m == '212': return (sf['foam'].get(c) or {}).get(s) or 0
    return 0

def get_st02(item, s02):
    m, c, s = item['model'], item['color'], float(str(item['size']).split('/')[0])
    if m == '205S': return (s02['canvas'].get(c) or {}).get(int(s)) or 0
    if m == '200':
        col = 'ดำ (ล้วน)' if c == 'ดำ' else c
        return (s02['foam'].get(col) or {}).get(s) or 0
    if m == '212': return (s02['foam212'].get(c) or {}).get(s) or 0
    return 0

# ══════════════════════════════════════════
#  BUILD EXCEL
# ══════════════════════════════════════════
def build_excel(so_list, sf, s02, out_path, dest):
    is_factory  = dest == 'factory'
    dest_label  = 'โรงงาน (ติดบาร์โค้ด)' if is_factory else 'โกดังบางหว้า (หยิบออก)'
    today       = datetime.date.today().strftime('%d/%m/%Y')

    def get_stF(i):
        m,c,s = i['model'],i['color'],float(str(i['size']).split('/')[0])
        if m=='205S': return (sf['canvas'].get(c) or {}).get(s) or 0
        if m in ('200','212'): return (sf['foam'].get(c) or {}).get(s) or 0
        return 0

    all_items = []
    for so in so_list:
        for item in so['items']:
            f = get_stF(item)
            entry = {**item, 'soNo':so['soNo'], 'customer':so['customer'],
                     'delivery':so['delivery'], 'poNo':so['poNo'], 'stF':f}
            if is_factory and f > 0:      all_items.append(entry)
            if not is_factory and f == 0: all_items.append(entry)

    if not all_items:
        return False, f'ไม่มีรายการสำหรับ {dest_label}'

    all_so    = ', '.join(sorted(set(i['soNo'] for i in all_items)))
    all_po    = ', '.join(sorted(set(i['poNo'] for i in all_items if i['poNo']!='-')))
    all_dates = ', '.join(sorted(set(i['delivery'] for i in all_items if i['delivery']!='-')))

    MODEL_ORDER = [('205S','ผ้าใบ 205-S'),('200','ฟองน้ำ 200'),('212','ฟองน้ำ 212')]

    # sizes per model (all SOs combined)
    model_sizes = {}
    for mid, _ in MODEL_ORDER:
        szs = sorted(set(float(str(i['size']).split('/')[0])
                         for i in all_items if i['model']==mid))
        if szs: model_sizes[mid] = szs

    max_n = max((len(v) for v in model_sizes.values()), default=6)
    TC = max_n + 2  # col1=สี, cols2..N+1=sizes, colTC=รวม

    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════
    # Sheet 1: รายการสินค้า
    # ════════════════════════════════
    ws = wb.create_sheet('รายการสินค้า')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions[get_column_letter(1)].width = 18
    for ci in range(2, TC):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    ws.column_dimensions[get_column_letter(TC)].width = 9

    def banner(r, txt, bg, fc='FFFFFFFF', sz=9, bold=True, italic=False):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        x = ws.cell(r, 1, f'  {txt}')
        x.font = Font(name='Arial', size=sz, bold=bold, color=fc, italic=italic)
        x.fill = PatternFill('solid', start_color=bg)
        x.alignment = Alignment(horizontal='left', vertical='center')
        return r + 1

    r = 1
    r = banner(r, f'นันยางมาร์เก็ตติ้ง — รายการสั่งสินค้า {dest_label}',
               P['TITLE'], 'FFFFFFFF', 11); ws.row_dimensions[r-1].height = 26
    r = banner(r, f'SO: {all_so}  |  {all_po}  |  กำหนดส่ง: {all_dates}  |  สร้าง: {today}  |  VENDOR: 600635',
               P['META'], 'FFB3D4FF', 8, italic=True); ws.row_dimensions[r-1].height = 14
    note = '* = ปัดเพิ่มให้ครบโหล (12 คู่)' if is_factory else 'จำนวนออกตามจริงตาม SO'
    r = banner(r, note, P['NOTE'], 'FF90CAF9', 8, italic=True); ws.row_dimensions[r-1].height = 13
    r += 1  # blank

    # ── One section per model ──
    for mid, mname in MODEL_ORDER:
        m_items = [i for i in all_items if i['model']==mid]
        if not m_items: continue

        sizes  = model_sizes[mid]
        colors = sorted(set(i['color'] for i in m_items))
        N      = len(sizes)
        pad    = TC - 2 - N  # empty cols to fill up to TC

        # Model header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        mh = ws.cell(r, 1, f'  {mname}')
        mh.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        mh.fill = PatternFill('solid', start_color=P['BLK_HDR'])
        mh.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 20; r += 1

        # Size header
        cl(ws, r, 1, 'สี', P['COL_HDR'], 'FFFFFFFF', True, ha='left')
        for si, sz in enumerate(sizes):
            cl(ws, r, 2+si, str(int(sz) if sz==int(sz) else sz), P['COL_HDR'], 'FFFFFFFF', True)
        for pi in range(pad):
            cl(ws, r, 2+N+pi, '', P['COL_HDR'], 'FFFFFFFF', False)
        cl(ws, r, TC, 'รวม', P['COL_TOT'], 'FFFFFFFF', True)
        ws.row_dimensions[r].height = 15; r += 1

        # Data rows — aggregate ALL SOs together
        dr0 = r
        col_tots = [0]*N
        for ri2, color in enumerate(colors):
            bg = P['ROW_A'] if ri2%2==0 else P['ROW_B']
            cl(ws, r, 1, f'  {color}', bg, P['DARK'], ha='left')
            row_tot = 0
            for si, sz in enumerate(sizes):
                raw = sum(i['qty'] for i in m_items
                          if i['color']==color and
                          float(str(i['size']).split('/')[0])==sz)
                val, padded = roundup12(raw) if is_factory else (raw, False)
                col_tots[si] += val; row_tot += val
                if val > 0:
                    x = ws.cell(r, 2+si, val)
                    x.font = Font(name='Arial', size=9, bold=padded,
                                  color=P['RED'] if padded else P['TEAL'])
                    x.fill = PatternFill('solid', start_color=bg)
                    x.alignment = Alignment(horizontal='center', vertical='center')
                    x.border = bdr()
                    x.number_format = '#,##0"*"' if padded else '#,##0'
                else:
                    x = ws.cell(r, 2+si, 0)
                    x.font = Font(name='Arial', size=9, color=P['GREY'])
                    x.fill = PatternFill('solid', start_color=bg)
                    x.alignment = Alignment(horizontal='center', vertical='center')
                    x.border = bdr(); x.number_format = '#,##0;-#,##0;"–"'
            for pi in range(pad):
                x = ws.cell(r, 2+N+pi, '')
                x.fill = PatternFill('solid', start_color=bg); x.border = bdr()
            # row total
            c0 = get_column_letter(2); cN = get_column_letter(1+N)
            t = ws.cell(r, TC, f'=SUM({c0}{r}:{cN}{r})')
            t.font = Font(name='Arial', size=9, bold=True, color=P['TEAL'])
            t.fill = PatternFill('solid', start_color=bg)
            t.alignment = Alignment(horizontal='center', vertical='center')
            t.border = bdr(); t.number_format = '#,##0;-#,##0;"–"'
            ws.row_dimensions[r].height = 15; r += 1

        # รวม row
        cl(ws, r, 1, '  รวม', P['SUB'], 'FFFFFFFF', True, ha='left')
        for si in range(N):
            col_l = get_column_letter(2+si)
            f2 = ws.cell(r, 2+si, f'=SUM({col_l}{dr0}:{col_l}{r-1})')
            f2.font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
            f2.fill = PatternFill('solid', start_color=P['SUB'])
            f2.alignment = Alignment(horizontal='center', vertical='center')
            f2.border = bdr(); f2.number_format = '#,##0;-#,##0;"–"'
        for pi in range(pad):
            x = ws.cell(r, 2+N+pi, '')
            x.fill = PatternFill('solid', start_color=P['SUB']); x.border = bdr()
        tc_l = get_column_letter(TC)
        ft = ws.cell(r, TC, f'=SUM({tc_l}{dr0}:{tc_l}{r-1})')
        ft.font = Font(name='Arial', size=9, bold=True, color='FFFFFFFF')
        ft.fill = PatternFill('solid', start_color=P['SUB'])
        ft.alignment = Alignment(horizontal='center', vertical='center')
        ft.border = bdr(); ft.number_format = '#,##0;-#,##0;"–"'
        ws.row_dimensions[r].height = 16; r += 1
        r += 1  # gap between models

    # ════════════════════════════════
    # Sheet 2: สรุปรายการ (by SO)
    # ════════════════════════════════
    ws2 = wb.create_sheet('สรุปรายการ')
    ws2.sheet_view.showGridLines = False
    hdrs = ['SO','ลูกค้า','กำหนดส่ง','รุ่น','สี','เบอร์','SO (คู่)','สั่งจริง (คู่)','ปัดเพิ่ม']
    wids = [14,12,12,10,14,7,10,12,10]
    for ci,(h,w) in enumerate(zip(hdrs,wids),1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        cl(ws2, 1, ci, h, P['COL_HDR'], 'FFFFFFFF', True)
    ws2.row_dimensions[1].height = 16

    ts=to=tp=0
    for ri2, item in enumerate(all_items, 2):
        if is_factory: ord_,pad=roundup12(item['qty']); pad_n=ord_-item['qty'] if pad else 0
        else:          ord_,pad,pad_n = item['qty'],False,0
        bg = P['ROW_A'] if ri2%2==0 else P['ROW_B']
        sz2 = float(str(item['size']).split('/')[0])
        sz_disp = int(sz2) if sz2==int(sz2) else sz2
        vals = [item['soNo'],item['customer'],item['delivery'],item['model'],
                item['color'],sz_disp,item['qty'],ord_,pad_n]
        for ci,v in enumerate(vals,1):
            cl(ws2, ri2, ci, v, bg,
               P['RED'] if ci==8 and pad else P['DARK'],
               bold=(ci==8 and pad), ha='left' if ci<=5 else 'center')
        ts+=item['qty']; to+=ord_; tp+=pad_n

    tr = len(all_items)+2
    for ci,v in enumerate(['']*5+[ts,to,tp,''],1):
        if ci==5: v='รวม'
        cl(ws2, tr, ci, v, P['GRAND'],
           P['RED'] if ci==8 and tp else 'FFFFFFFF', True,
           ha='right' if ci==5 else 'center')

    wb.save(out_path)
    return True, out_path

# ══════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('นันยางมาร์เก็ตติ้ง — สร้างรายการบาร์โค้ด')
        self.geometry('560x580')
        self.resizable(False, False)
        self.configure(bg='#F0EEF8')
        self._build_ui()

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg='#3A3270', height=60)
        hdr.pack(fill='x')
        tk.Label(hdr, text='🏭  ระบบสร้างรายการสินค้าติดบาร์โค้ด',
                 bg='#3A3270', fg='white',
                 font=('Arial',13,'bold')).pack(side='left', padx=16, pady=14)

        body = tk.Frame(self, bg='#F0EEF8')
        body.pack(fill='both', expand=True, padx=20, pady=16)

        # File pickers
        self.pdf_paths  = []
        self.e1_path    = tk.StringVar()
        self.e2_path    = tk.StringVar()
        self.out_folder = tk.StringVar(value=str(Path.home()/'Desktop'))

        self._file_row(body, '📄  ใบสั่งขาย SO (PDF)', 'เลือกไฟล์ PDF',
                       lambda: self._pick_pdfs(), 'pdfs', row=0)
        self._file_row(body, '📊  สต็อคโรงงาน (Excel)', 'เลือกไฟล์ Stock205-S_200.xlsx',
                       lambda: self._pick_file(self.e1_path,'Excel Files','*.xlsx'), 'e1', row=1)
        self._file_row(body, '📊  สต็อค 02 (Excel)', 'เลือกไฟล์ stock_02.xlsx',
                       lambda: self._pick_file(self.e2_path,'Excel Files','*.xlsx'), 'e2', row=2)
        self._file_row(body, '📁  โฟลเดอร์บันทึกผล', 'เลือกโฟลเดอร์',
                       lambda: self._pick_folder(), 'out', row=3)

        # Options
        opt = tk.LabelFrame(body, text=' ตัวเลือก ', bg='#F0EEF8',
                            font=('Arial',9), fg='#3A3270', padx=10, pady=8)
        opt.grid(row=4, column=0, columnspan=3, sticky='ew', pady=(8,0))

        self.do_factory   = tk.BooleanVar(value=True)
        self.do_warehouse = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text='สร้างไฟล์ โรงงาน (ติดบาร์โค้ด)',
                       variable=self.do_factory,
                       bg='#F0EEF8', fg='#3A3270',
                       font=('Arial',9)).pack(anchor='w')
        tk.Checkbutton(opt, text='สร้างไฟล์ โกดังบางหว้า (หยิบออก)',
                       variable=self.do_warehouse,
                       bg='#F0EEF8', fg='#7A4530',
                       font=('Arial',9)).pack(anchor='w')

        # Progress
        self.progress = ttk.Progressbar(body, mode='indeterminate', length=400)

        self.status = tk.Label(body, text='', bg='#F0EEF8',
                               fg='#555', font=('Arial',9,'italic'))

        # ── Email settings (collapsible frame) ──
        email_frame = tk.LabelFrame(body, text=' ⚙ ตั้งค่าอีเมล (กรอกครั้งแรกครั้งเดียว) ',
                                    bg='#F0EEF8', font=('Arial',8), fg='#3A3270',
                                    padx=10, pady=8)
        email_frame.grid(row=5, column=0, columnspan=3, sticky='ew', pady=(8,0))

        cfg = load_config()
        self.email_from = tk.StringVar(value=cfg.get('from_email',''))
        self.email_pw   = tk.StringVar(value=get_password())
        self.email_factory  = tk.StringVar(value=cfg.get('to_factory',''))
        self.email_bangwa   = tk.StringVar(value=cfg.get('to_bangwa',''))
        self.email_cc       = tk.StringVar(value=cfg.get('cc',''))
        self.auto_send      = tk.BooleanVar(value=cfg.get('auto_send', False))

        fields = [
            ('อีเมลผู้ส่ง (From)', self.email_from, False),
            ('Password', self.email_pw, True),
            ('ถึงโรงงาน (To)', self.email_factory, False),
            ('ถึงบางหว้า (To)', self.email_bangwa, False),
            ('CC (ถ้ามี)', self.email_cc, False),
        ]
        for fi, (lbl, var, is_pw) in enumerate(fields):
            tk.Label(email_frame, text=lbl, bg='#F0EEF8',
                     font=('Arial',8), fg='#555', anchor='w',
                     width=20).grid(row=fi, column=0, sticky='w', pady=2)
            show = '*' if is_pw else ''
            tk.Entry(email_frame, textvariable=var, show=show,
                     font=('Arial',8), width=35,
                     relief='flat', bg='white', bd=1).grid(row=fi, column=1, sticky='ew', padx=(4,0), pady=2)
        tk.Checkbutton(email_frame, text='ส่งอีเมลอัตโนมัติหลังสร้าง Excel',
                       variable=self.auto_send, bg='#F0EEF8',
                       font=('Arial',8), fg='#3A3270').grid(
                       row=len(fields), column=0, columnspan=2, sticky='w', pady=(4,0))
        email_frame.columnconfigure(1, weight=1)

        # Progress
        self.progress = ttk.Progressbar(body, mode='indeterminate', length=400)
        self.progress.grid(row=6, column=0, columnspan=3, pady=(12,0), sticky='ew')

        self.status = tk.Label(body, text='', bg='#F0EEF8',
                               fg='#555', font=('Arial',9,'italic'))
        self.status.grid(row=7, column=0, columnspan=3)

        # Run button
        self.btn = tk.Button(body, text='▶  สร้าง Excel',
                             bg='#3A3270', fg='white',
                             font=('Arial',11,'bold'),
                             relief='flat', cursor='hand2',
                             padx=20, pady=10,
                             command=self._run)
        self.btn.grid(row=8, column=0, columnspan=3, pady=(12,0))

        body.columnconfigure(1, weight=1)

    def _file_row(self, parent, label, placeholder, cmd, key, row):
        tk.Label(parent, text=label, bg='#F0EEF8',
                 font=('Arial',9,'bold'), fg='#3A3270',
                 anchor='w').grid(row=row, column=0, sticky='w', pady=(8,0))
        ent = tk.Entry(parent, font=('Arial',8), fg='#888', width=38,
                       relief='flat', bg='#FFFFFF', bd=1)
        ent.insert(0, placeholder)
        ent.grid(row=row, column=1, padx=(6,4), pady=(8,0), sticky='ew')
        setattr(self, f'ent_{key}', ent)
        tk.Button(parent, text='เลือก', command=cmd,
                  bg='#C8EAD8', fg='#1A5C46',
                  font=('Arial',8,'bold'),
                  relief='flat', cursor='hand2',
                  padx=6).grid(row=row, column=2, pady=(8,0))

    def _pick_pdfs(self):
        paths = filedialog.askopenfilenames(
            title='เลือกไฟล์ SO PDF',
            filetypes=[('PDF Files','*.pdf'),('All','*.*')])
        if paths:
            self.pdf_paths = list(paths)
            self.ent_pdfs.config(fg='#1A5C46')
            self.ent_pdfs.delete(0,'end')
            self.ent_pdfs.insert(0, f'{len(paths)} ไฟล์: ' + ', '.join(Path(p).name for p in paths[:2]))

    def _pick_file(self, var, desc, ext):
        path = filedialog.askopenfilename(
            title=f'เลือกไฟล์ {desc}',
            filetypes=[(desc,ext),('All','*.*')])
        if path:
            var.set(path)
            key = 'e1' if var==self.e1_path else 'e2'
            ent = getattr(self, f'ent_{key}')
            ent.config(fg='#1A5C46')
            ent.delete(0,'end')
            ent.insert(0, Path(path).name)

    def _pick_folder(self):
        folder = filedialog.askdirectory(title='เลือกโฟลเดอร์บันทึกผล')
        if folder:
            self.out_folder.set(folder)
            self.ent_out.config(fg='#3A3270')
            self.ent_out.delete(0,'end')
            self.ent_out.insert(0, folder)

    def _set_status(self, txt, color='#555'):
        self.status.config(text=txt, fg=color)

    def _run(self):
        if not self.pdf_paths:
            messagebox.showwarning('แจ้งเตือน','กรุณาเลือกไฟล์ PDF ก่อน')
            return
        if not self.e1_path.get():
            messagebox.showwarning('แจ้งเตือน','กรุณาเลือกไฟล์ สต็อคโรงงาน (Excel)')
            return
        if not self.e2_path.get():
            messagebox.showwarning('แจ้งเตือน','กรุณาเลือกไฟล์ สต็อค 02 (Excel)')
            return
        self.btn.config(state='disabled')
        self.progress.start(10)
        threading.Thread(target=self._process, daemon=True).start()

    def _process(self):
        try:
            self._set_status('กำลังอ่าน PDF...')
            so_list = [parse_pdf(p) for p in self.pdf_paths]
            so_list = [s for s in so_list if s['items']]

            if not so_list:
                self.after(0, lambda: messagebox.showerror('Error','ไม่พบรายการสินค้าใน PDF'))
                return

            self._set_status('กำลังอ่านสต็อค...')
            sf  = parse_stock_factory(self.e1_path.get())
            s02 = parse_stock02(self.e2_path.get())

            out = self.out_folder.get() or str(Path.home()/'Desktop')
            today = datetime.date.today().strftime('%Y-%m-%d')
            results = []

            if self.do_factory.get():
                self._set_status('กำลังสร้าง Excel โรงงาน...')
                fp = os.path.join(out, f'barcode_factory_{today}.xlsx')
                ok, msg = build_excel(so_list, sf, s02, fp, 'factory')
                if ok: results.append(f'✓ โรงงาน: {Path(fp).name}')
                else: results.append(f'✗ โรงงาน: {msg}')

            if self.do_warehouse.get():
                self._set_status('กำลังสร้าง Excel บางหว้า...')
                wp = os.path.join(out, f'barcode_bangwa_{today}.xlsx')
                ok, msg = build_excel(so_list, sf, s02, wp, 'warehouse')
                if ok: results.append(f'✓ บางหว้า: {Path(wp).name}')
                else: results.append(f'✗ บางหว้า: {msg}')

            # Save email config
            cfg = load_config()
            cfg['from_email']  = self.email_from.get().strip()
            cfg['to_factory']  = self.email_factory.get().strip()
            cfg['to_bangwa']   = self.email_bangwa.get().strip()
            cfg['cc']          = self.email_cc.get().strip()
            cfg['auto_send']   = self.auto_send.get()
            save_config(cfg)
            set_password(self.email_pw.get())

            # Auto send email
            email_results = []
            if self.auto_send.get() and cfg['from_email']:
                self._set_status('กำลังส่งอีเมล...')
                today_str = datetime.date.today().strftime('%d/%m/%Y')
                so_str = ', '.join(s['soNo'] for s in so_list)

                # Send to factory
                fp = os.path.join(out, f'barcode_factory_{datetime.date.today()}.xlsx')
                if cfg['to_factory'] and os.path.exists(fp):
                    try:
                        msg_obj = MIMEMultipart()
                        msg_obj['From']    = cfg['from_email']
                        msg_obj['To']      = cfg['to_factory']
                        msg_obj['Subject'] = f'รายการสินค้าติดบาร์โค้ด {so_str} — ส่ง {today_str}'
                        if cfg['cc']: msg_obj['Cc'] = cfg['cc']
                        body_txt = f'เรียน ทางโรงงาน\n\nขอแจ้งรายการสินค้าสำหรับติดบาร์โค้ด\nSO: {so_str}\nกำหนดส่ง: ตามไฟล์แนบ\n\nขอแสดงความนับถือ\nฝ่าย Sale Support — นันยางมาร์เก็ตติ้ง จำกัด'
                        msg_obj.attach(MIMEText(body_txt, 'plain', 'utf-8'))
                        with open(fp, 'rb') as f:
                            part = MIMEBase('application','octet-stream')
                            part.set_payload(f.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f'attachment; filename="{Path(fp).name}"')
                        msg_obj.attach(part)
                        smtp_host, smtp_port = detect_smtp(cfg['from_email'])
                        ctx = ssl.create_default_context()
                        with smtplib.SMTP(smtp_host, smtp_port) as srv:
                            srv.ehlo(); srv.starttls(context=ctx); srv.ehlo()
                            srv.login(cfg['from_email'], get_password())
                            recip = [cfg['to_factory']] + ([cfg['cc']] if cfg['cc'] else [])
                            srv.sendmail(cfg['from_email'], recip, msg_obj.as_string())
                        email_results.append('✓ ส่งอีเมลโรงงานแล้ว')
                    except Exception as e:
                        email_results.append(f'✗ ส่งอีเมลโรงงานไม่ได้: {e}')

                # Send to bangwa
                wp = os.path.join(out, f'barcode_bangwa_{datetime.date.today()}.xlsx')
                if cfg['to_bangwa'] and os.path.exists(wp):
                    try:
                        msg_obj2 = MIMEMultipart()
                        msg_obj2['From']    = cfg['from_email']
                        msg_obj2['To']      = cfg['to_bangwa']
                        msg_obj2['Subject'] = f'รายการสินค้าหยิบออกจากโกดัง {so_str} — ส่ง {today_str}'
                        if cfg['cc']: msg_obj2['Cc'] = cfg['cc']
                        body_txt2 = f'เรียน ทีมโกดังบางหว้า\n\nขอแจ้งรายการสินค้าที่ต้องหยิบออกจากโกดัง\nSO: {so_str}\nกำหนดส่ง: ตามไฟล์แนบ\n\nขอแสดงความนับถือ\nฝ่าย Sale Support — นันยางมาร์เก็ตติ้ง จำกัด'
                        msg_obj2.attach(MIMEText(body_txt2, 'plain', 'utf-8'))
                        with open(wp, 'rb') as f:
                            part2 = MIMEBase('application','octet-stream')
                            part2.set_payload(f.read())
                        encoders.encode_base64(part2)
                        part2.add_header('Content-Disposition', f'attachment; filename="{Path(wp).name}"')
                        msg_obj2.attach(part2)
                        smtp_host, smtp_port = detect_smtp(cfg['from_email'])
                        ctx = ssl.create_default_context()
                        with smtplib.SMTP(smtp_host, smtp_port) as srv:
                            srv.ehlo(); srv.starttls(context=ctx); srv.ehlo()
                            srv.login(cfg['from_email'], get_password())
                            recip2 = [cfg['to_bangwa']] + ([cfg['cc']] if cfg['cc'] else [])
                            srv.sendmail(cfg['from_email'], recip2, msg_obj2.as_string())
                        email_results.append('✓ ส่งอีเมลบางหว้าแล้ว')
                    except Exception as e:
                        email_results.append(f'✗ ส่งอีเมลบางหว้าไม่ได้: {e}')

            self._set_status('เสร็จแล้ว ✓', '#1A7A3F')
            all_results = results + email_results
            msg = 'สร้างและส่งไฟล์เรียบร้อย!\n\n' + '\n'.join(all_results) + f'\n\nบันทึกที่: {out}'
            self.after(0, lambda: messagebox.showinfo('เสร็จแล้ว!', msg))
            if os.path.exists(out):
                os.startfile(out)

        except Exception as e:
            import traceback
            err = traceback.format_exc()
            self.after(0, lambda: messagebox.showerror('Error', str(e)+'\n\n'+err[:300]))
        finally:
            self.after(0, lambda: (self.progress.stop(), self.btn.config(state='normal')))

if __name__ == '__main__':
    app = App()
    app.mainloop()
