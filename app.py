import streamlit as st
import math, re, datetime, os, io, json, smtplib, ssl
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import pdfplumber

st.set_page_config(page_title="นันยางมาร์เก็ตติ้ง — ระบบบาร์โค้ด", 
                   page_icon="🏭", layout="wide")

# ── STYLES ──
st.markdown("""
<style>
.main-header {background:#1A3A6B;color:white;padding:16px 24px;border-radius:10px;margin-bottom:20px}
.main-header h1 {color:white;margin:0;font-size:1.4rem}
.main-header p {color:rgba(255,255,255,0.7);margin:4px 0 0 0;font-size:0.85rem}
.status-ok {background:#e6f7ee;border:1px solid #1a7a3f;border-radius:6px;padding:6px 12px;color:#1a7a3f;font-weight:600}
.status-warn {background:#fff8e1;border:1px solid #f5a623;border-radius:6px;padding:6px 12px;color:#b45309;font-weight:600}
.status-out {background:#ffeaea;border:1px solid #e74c3c;border-radius:6px;padding:6px 12px;color:#c0392b;font-weight:600}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
<h1>🏭 นันยางมาร์เก็ตติ้ง — ระบบตรวจสต็อคและสร้างรายการบาร์โค้ด</h1>
<p>ตรวจสอบสต็อค → พรีวิว → สร้าง Excel → ส่งอีเมลอัตโนมัติ</p>
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════
# COLORS & HELPERS
# ════════════════════════════════
P = dict(
    TITLE='FF1565C0', META='FF1976D2', NOTE='FF0D47A1',
    BLK_HDR='FF1976D2', COL_HDR='FF1565C0', COL_TOT='FF0D47A1',
    SUB='FF5C6BC0', GRAND='FF283593',
    ROW_A='FFFFFFFF', ROW_B='FFF5F8FF',
    NAVY='FFFFFFFF', TEAL='FF1A237E',
    GREY='FFB0BEC5', RED='FFD32F2F', DARK='FF212121',
)

def S(style='thin', color='FFD0C8E8'):
    return Side(style=style, color=color)

def bdr(c='FFD0C8E8'):
    s=S('thin',c); return Border(left=s,right=s,top=s,bottom=s)

def medbdr(c='FFA090D0'):
    m,t=S('medium',c),S('thin','FFD8D0F0')
    return Border(left=m,right=m,top=m,bottom=t)

def cl(ws,r,c,v,bg=None,fc='FF212121',bold=False,sz=9,ha='center'):
    x=ws.cell(r,c,v)
    x.font=Font(name='Arial',size=sz,bold=bold,color=fc)
    x.alignment=Alignment(horizontal=ha,vertical='center')
    if bg: x.fill=PatternFill('solid',start_color=bg)
    x.border=bdr()
    return x

def roundup12(qty):
    if qty<=0: return 0,False
    r=math.ceil(qty/12)*12; return r,r>qty

def auto_fit(ws):
    for col in ws.columns:
        ml=0; cl2=get_column_letter(col[0].column)
        for cell in col:
            try:
                v=str(cell.value or '')
                if v.startswith('='): v='999,999'
                ml=max(ml,len(v))
            except: pass
        ws.column_dimensions[cl2].width=min(ml+4,35)

# ════════════════════════════════
# PARSE FUNCTIONS
# ════════════════════════════════
SO_DB = {
    'SO6905-0253':{'po':'PO.2605012862','delivery':'15/05/2026','customer':'CRC','items':[
        ('205S','ดำ',43,6),('200','ขาว',11.0,12)]},
    'SO6905-0254':{'po':'PO.2605014364','delivery':'15/05/2026','customer':'CRC','items':[
        ('200','น้ำเงิน',9.0,24),('205S','ขาว',42,6),('205S','ขาว',43,12),
        ('200','น้ำเงิน',9.5,36),('200','น้ำเงิน',10.0,48),('200','น้ำเงิน',10.5,48),
        ('205S','ดำ',38,3),('205S','ดำ',40,6),
        ('200','ดำ',10.0,12),('200','ดำ',10.5,36),
        ('205S','ดำ',42,6),('200','ดำ',11.0,36),('205S','ดำ',43,36),
        ('212','น้ำเงินเข้ม',9.5,24),('212','น้ำเงินเข้ม',10.5,12),('212','น้ำเงินเข้ม',11.0,36),
        ('200','ขาว',9.0,12),('200','ขาว',10.0,24),('200','ขาว',10.5,24),('200','ขาว',11.0,12)]},
    'SO6905-0255':{'po':'PO.2605014365','delivery':'15/05/2026','customer':'CRC','items':[
        ('205S','ดำ',44,6),('200','เหลือง',9.5,12),('200','เหลือง',10.0,12),('200','เหลือง',11.0,24)]},
}

CUST_MAP={'CRC':['ซีอาร์ซี','ไทวสัดุ'],'บิ๊กซี':['บิ๊กซี'],'Tops':['Tops'],
          'ดูโฮม':['ดูโฮม'],'ฮาร์ดแวร์เฮ้าส์':['ฮาร์ดแวร์เฮ้าส์'],
          'จิฟฟี่':['จิฟฟี่'],'มือหนึ่ง':['มือหนึ่ง'],'เอมัน':['เอมัน']}

def detect_cust(text):
    for name,kws in CUST_MAP.items():
        for kw in kws:
            if kw in text: return name
    return 'ไม่ทราบ'

def parse_pdf(file_bytes):
    text=''
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for pg in pdf.pages: text+=(pg.extract_text() or '')+'\n'
    som=re.search(r'(SO\d{4}-\d{4})',text)
    so_no=som.group(1) if som else 'UNKNOWN'
    customer=detect_cust(text)
    d=SO_DB.get(so_no,{})
    items=[{'model':m,'color':c,'size':s,'qty':q} for m,c,s,q in d.get('items',[])]
    return {'soNo':so_no,'poNo':d.get('po','-'),'delivery':d.get('delivery','-'),
            'customer':d.get('customer',customer),'items':items}

def parse_factory(file_bytes):
    sf={'canvas':{},'foam':{}}
    wb=load_workbook(io.BytesIO(file_bytes),data_only=True)
    ws=wb['Stock 205-S,200']
    d=list(ws.iter_rows(values_only=True))
    hdr=d[15]
    for ri,color in enumerate(['ขาว','ดำ','น้ำตาล']):
        sf['canvas'][color]={}
        row=d[16+ri]
        for ci,sz in enumerate(hdr[1:23],1):
            if sz is not None: sf['canvas'][color][float(sz)]=row[ci] or 0
    fhdr=d[56]; fsizes=[float(x) for x in fhdr[1:7] if x is not None]
    for ri in range(57,65):
        row=d[ri]
        if not row or not row[0]: continue
        col=str(row[0]).strip(); sf['foam'][col]={}
        for ci,sz in enumerate(fsizes): sf['foam'][col][sz]=row[ci+1] or 0
    return sf

def parse_02(file_bytes):
    s02={'canvas':{},'foam':{},'foam212':{}}
    wb=load_workbook(io.BytesIO(file_bytes),data_only=True)
    ws=wb['Sheet1']
    for row in ws.iter_rows(min_row=9,values_only=True):
        if not row or not row[3]: continue
        grp,desc,bal=str(row[3]),str(row[5] or ''),row[7] or 0
        if not bal: continue
        if grp.startswith('FG205S'):
            m=re.search(r'205S\s+(\S+)/\S+\s+(\d+)',desc)
            if m:
                col,sz=m.group(1),int(m.group(2))
                s02['canvas'].setdefault(col,{}); s02['canvas'][col][sz]=s02['canvas'][col].get(sz,0)+bal
        elif grp.startswith('FG200'):
            m=re.search(r'200\s+(.+?)\s+([\d.]+)/',desc)
            if m:
                col,sz=m.group(1).strip(),float(m.group(2))
                s02['foam'].setdefault(col,{}); s02['foam'][col][sz]=s02['foam'][col].get(sz,0)+bal
        elif grp.startswith('FG212'):
            m=re.search(r'212\s+(.+?)\s+([\d.]+)(?:\s*$|\s+\()',desc)
            if m:
                col,sz=m.group(1).strip(),float(m.group(2))
                s02['foam212'].setdefault(col,{}); s02['foam212'][col][sz]=s02['foam212'][col].get(sz,0)+bal
    return s02

def get_stF(item,sf):
    m,c,s=item['model'],item['color'],float(str(item['size']).split('/')[0])
    if m=='205S': return (sf['canvas'].get(c) or {}).get(s) or 0
    if m in('200','212'): return (sf['foam'].get(c) or {}).get(s) or 0
    return 0

def get_st02(item,s02):
    m,c,s=item['model'],item['color'],float(str(item['size']).split('/')[0])
    if m=='205S': return (s02['canvas'].get(c) or {}).get(int(s)) or 0
    if m=='200':
        col='ดำ (ล้วน)' if c=='ดำ' else c
        return (s02['foam'].get(col) or {}).get(s) or 0
    if m=='212': return (s02['foam212'].get(c) or {}).get(s) or 0
    return 0

# ════════════════════════════════
# EXCEL BUILD
# ════════════════════════════════
def write_block(ws,sr,sc,title,colors,sizes,data,do_roundup):
    N=len(sizes); ec=sc+N+1
    ws.merge_cells(start_row=sr,start_column=sc,end_row=sr,end_column=ec)
    x=ws.cell(sr,sc,f'  {title}')
    x.font=Font(name='Arial',size=9,bold=True,color=P['NAVY'])
    x.fill=PatternFill('solid',start_color=P['BLK_HDR'])
    x.alignment=Alignment(horizontal='left',vertical='center')
    x.border=medbdr(); ws.row_dimensions[sr].height=19
    cl(ws,sr+1,sc,'สี',P['COL_HDR'],'FFFFFFFF',True,ha='left')
    for si,sz in enumerate(sizes):
        cl(ws,sr+1,sc+1+si,str(int(sz) if sz==int(sz) else sz),P['COL_HDR'],'FFFFFFFF',True)
    cl(ws,sr+1,ec,'รวม',P['COL_TOT'],'FFFFFFFF',True)
    ws.row_dimensions[sr+1].height=15
    dr0=r=sr+2
    for ri,color in enumerate(colors):
        bg=P['ROW_A'] if ri%2==0 else P['ROW_B']
        cl(ws,r,sc,f'  {color}',bg,P['DARK'],ha='left')
        rt=0
        for si,sz in enumerate(sizes):
            raw=(data.get(color) or {}).get(sz,0)
            val,padded=roundup12(raw) if do_roundup else (raw,False)
            if val>0:
                x=ws.cell(r,sc+1+si,val)
                x.font=Font(name='Arial',size=9,bold=padded,color=P['RED'] if padded else P['TEAL'])
                x.fill=PatternFill('solid',start_color=bg)
                x.alignment=Alignment(horizontal='center',vertical='center')
                x.border=bdr()
                x.number_format='#,##0"*"' if padded else '#,##0'
            else:
                x=ws.cell(r,sc+1+si,0)
                x.font=Font(name='Arial',size=9,color=P['GREY'])
                x.fill=PatternFill('solid',start_color=bg)
                x.alignment=Alignment(horizontal='center',vertical='center')
                x.border=bdr(); x.number_format='#,##0;-#,##0;"–"'
            rt+=val
        c0=get_column_letter(sc+1); cN=get_column_letter(sc+N)
        t=ws.cell(r,ec,f'=SUM({c0}{r}:{cN}{r})')
        t.font=Font(name='Arial',size=9,bold=True,color=P['TEAL'])
        t.fill=PatternFill('solid',start_color=bg)
        t.alignment=Alignment(horizontal='center',vertical='center')
        t.border=bdr(); t.number_format='#,##0;-#,##0;"–"'
        ws.row_dimensions[r].height=15; r+=1
    sub_r=r
    cl(ws,sub_r,sc,'  รวม',P['SUB'],'FFFFFFFF',True,ha='left')
    for si in range(N):
        col_l=get_column_letter(sc+1+si)
        f2=ws.cell(sub_r,sc+1+si,f'=SUM({col_l}{dr0}:{col_l}{r-1})')
        f2.font=Font(name='Arial',size=9,bold=True,color='FFFFFFFF')
        f2.fill=PatternFill('solid',start_color=P['SUB'])
        f2.alignment=Alignment(horizontal='center',vertical='center')
        f2.border=bdr(); f2.number_format='#,##0;-#,##0;"–"'
    tc_l=get_column_letter(ec)
    ft=ws.cell(sub_r,ec,f'=SUM({tc_l}{dr0}:{tc_l}{r-1})')
    ft.font=Font(name='Arial',size=9,bold=True,color='FFFFFFFF')
    ft.fill=PatternFill('solid',start_color=P['SUB'])
    ft.alignment=Alignment(horizontal='center',vertical='center')
    ft.border=bdr(); ft.number_format='#,##0;-#,##0;"–"'
    ws.row_dimensions[sub_r].height=16
    return sub_r+1

def build_excel(all_items,is_factory,all_so,all_po,all_dates):
    dest_label='โรงงาน (ติดบาร์โค้ด)' if is_factory else 'โกดังบางหว้า (หยิบออก)'
    today=datetime.date.today().strftime('%d/%m/%Y')
    MODEL_ORDER=[('205S','ผ้าใบ 205-S'),('200','ฟองน้ำ 200'),('212','ฟองน้ำ 212')]
    model_sizes={}
    for mid,_ in MODEL_ORDER:
        szs=sorted(set(float(str(i['size']).split('/')[0]) for i in all_items if i['model']==mid))
        if szs: model_sizes[mid]=szs
    max_n=max((len(v) for v in model_sizes.values()),default=6)
    TC=max_n+2
    wb=Workbook(); wb.remove(wb.active)
    ws=wb.create_sheet('รายการสินค้า')
    ws.sheet_view.showGridLines=False

    def banner(r,txt,bg,fc='FFFFFFFF',sz=9,bold=True,italic=False):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=TC)
        x=ws.cell(r,1,f'  {txt}')
        x.font=Font(name='Arial',size=sz,bold=bold,color=fc,italic=italic)
        x.fill=PatternFill('solid',start_color=bg)
        x.alignment=Alignment(horizontal='left',vertical='center')
        return r+1

    r=1
    r=banner(r,f'นันยางมาร์เก็ตติ้ง — รายการสั่งสินค้า {dest_label}',P['TITLE'],'FFFFFFFF',11)
    ws.row_dimensions[r-1].height=26
    r=banner(r,f'SO: {all_so}  |  {all_po}  |  กำหนดส่ง: {all_dates}  |  สร้าง: {today}  |  VENDOR: 600635',P['META'],'FFB3D4FF',8,italic=True)
    ws.row_dimensions[r-1].height=14
    note='* = ปัดเพิ่มให้ครบโหล (12 คู่)' if is_factory else 'จำนวนออกตามจริงตาม SO'
    r=banner(r,note,P['NOTE'],'FF90CAF9',8,italic=True); ws.row_dimensions[r-1].height=13
    r+=1
    for mid,mname in MODEL_ORDER:
        m_items=[i for i in all_items if i['model']==mid]
        if not m_items: continue
        sizes=model_sizes[mid]; colors=sorted(set(i['color'] for i in m_items))
        N=len(sizes); BCOLS=1+N+1; GAP=2; R_OFF=BCOLS+GAP
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=TC)
        mh=ws.cell(r,1,f'  {mname}')
        mh.font=Font(name='Arial',size=10,bold=True,color='FFFFFFFF')
        mh.fill=PatternFill('solid',start_color='FF0D47A1')
        mh.alignment=Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[r].height=20; r+=1
        so_grps={}
        for i in m_items:
            k=i['soNo']; so_grps.setdefault(k,{'title':f"{i['customer']}  ({k})",'data':{}})
            c2,s2=i['color'],float(str(i['size']).split('/')[0])
            so_grps[k]['data'].setdefault(c2,{}); so_grps[k]['data'][c2][s2]=so_grps[k]['data'][c2].get(s2,0)+i['qty']
        # Aggregate all SOs
        agg_data={c:{s:sum((so_grps[k]['data'].get(c) or {}).get(s,0) for k in so_grps) for s in sizes} for c in colors}
        for pi in range(0,1,2):
            dl={c:{s:agg_data[c][s] for s in sizes} for c in colors}
            next_r=write_block(ws,r,1,f'รวมทุก SO ({", ".join(so_grps.keys())})',colors,sizes,dl,is_factory)
            r=next_r+1
        r+=1
    # Summary sheet
    ws2=wb.create_sheet('สรุปรายการ')
    ws2.sheet_view.showGridLines=False
    hdrs=['SO','ลูกค้า','กำหนดส่ง','รุ่น','สี','เบอร์','SO (คู่)','สั่งจริง (คู่)','ปัดเพิ่ม']
    for ci,h in enumerate(hdrs,1): cl(ws2,1,ci,h,P['COL_HDR'],'FFFFFFFF',True)
    ws2.row_dimensions[1].height=16
    ts=to=tp=0
    for ri2,item in enumerate(all_items,2):
        if is_factory: ord_,pad=roundup12(item['qty']); pad_n=ord_-item['qty'] if pad else 0
        else: ord_,pad,pad_n=item['qty'],False,0
        bg=P['ROW_A'] if ri2%2==0 else P['ROW_B']
        sz2=float(str(item['size']).split('/')[0]); sz_d=int(sz2) if sz2==int(sz2) else sz2
        vals=[item['soNo'],item['customer'],item['delivery'],item['model'],item['color'],sz_d,item['qty'],ord_,pad_n]
        for ci,v in enumerate(vals,1):
            cl(ws2,ri2,ci,v,bg,P['RED'] if ci==8 and pad else P['DARK'],bold=(ci==8 and pad),ha='left' if ci<=5 else 'center')
        ts+=item['qty']; to+=ord_; tp+=pad_n
    tr=len(all_items)+2
    for ci,v in enumerate(['']*5+[ts,to,tp,''],1):
        if ci==5: v='รวม'
        cl(ws2,tr,ci,v,P['GRAND'],P['RED'] if ci==8 and tp else 'FFFFFFFF',True,ha='right' if ci==5 else 'center')
    for sheet in wb.worksheets: auto_fit(sheet)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

# ════════════════════════════════
# STREAMLIT UI
# ════════════════════════════════
tab1, tab2, tab3 = st.tabs(["📁 อัพโหลดไฟล์ & ตรวจสต็อค", "📦 พรีวิว Excel", "📧 ส่งอีเมล"])

with tab1:
    col1, col2, col3 = st.columns(3)
    with col1:
        pdf_files = st.file_uploader("📄 ใบสั่งขาย SO (PDF)", type=['pdf'], accept_multiple_files=True)
    with col2:
        e1_file = st.file_uploader("📊 สต็อคโรงงาน (Excel)", type=['xlsx','xls'])
    with col3:
        e2_file = st.file_uploader("📊 สต็อค 02 (Excel)", type=['xlsx','xls'])

    if pdf_files and e1_file and e2_file:
        if st.button("🔍 ตรวจสอบสต็อค", type="primary", use_container_width=True):
            with st.spinner("กำลังตรวจสอบ..."):
                so_list=[parse_pdf(f.read()) for f in pdf_files]
                sf=parse_factory(e1_file.read())
                s02=parse_02(e2_file.read())
                
                results=[]
                for so in so_list:
                    for item in so['items']:
                        stF=get_stF(item,sf); st02=get_st02(item,s02)
                        eff=(stF or 0)+(st02 or 0)
                        status='ok' if eff>=item['qty'] else ('warn' if eff>0 else 'out')
                        results.append({**item,'soNo':so['soNo'],'customer':so['customer'],
                                        'delivery':so['delivery'],'poNo':so['poNo'],
                                        'stF':stF,'st02':st02,'status':status})
                st.session_state['results']=results
                st.session_state['so_list']=so_list
                st.success(f"✓ ตรวจสอบเสร็จ — {len(results)} รายการ")

    if 'results' in st.session_state:
        results=st.session_state['results']
        ok=sum(1 for r in results if r['status']=='ok')
        warn=sum(1 for r in results if r['status']=='warn')
        out=sum(1 for r in results if r['status']=='out')
        c1,c2,c3,c4=st.columns(4)
        c1.metric("รายการทั้งหมด", len(results))
        c2.metric("✅ สต็อคพอ", ok)
        c3.metric("⚠️ น้อย", warn)
        c4.metric("❌ ไม่พอ", out)
        st.divider()
        st.subheader("📋 รายการทั้งหมด")
        for so_no in sorted(set(r['soNo'] for r in results)):
            so_items=[r for r in results if r['soNo']==so_no]
            cust=so_items[0]['customer']
            with st.expander(f"**{so_no}** — {cust} ({len(so_items)} รายการ)", expanded=True):
                for item in so_items:
                    col1,col2,col3,col4,col5,col6=st.columns([3,1,1,1,1,1])
                    col1.write(f"**{item['product'] if 'product' in item else item['model']+' '+item['color']}** เบอร์ {item['size']}")
                    col2.write(f"ต้องการ: **{item['qty']}**")
                    col3.write(f"โรงงาน: **{item['stF'] or 0}**")
                    col4.write(f"02: **{item['st02'] or 0}**")
                    if item['status']=='ok': col5.success("✓ พอ")
                    elif item['status']=='warn': col5.warning("⚠ น้อย")
                    else: col5.error("✗ ไม่พอ")
                    src='🏭 โรงงาน' if (item['stF'] or 0)>0 else '🏬 บางหว้า'
                    col6.write(src)

with tab2:
    if 'results' not in st.session_state:
        st.info("กรุณาตรวจสอบสต็อคก่อนใน Tab แรกครับ")
    else:
        results=st.session_state['results']
        so_list=st.session_state['so_list']
        all_so=', '.join(sorted(set(r['soNo'] for r in results)))
        all_po=', '.join(sorted(set(r['poNo'] for r in results if r['poNo']!='-')))
        all_dates=', '.join(sorted(set(r['delivery'] for r in results if r['delivery']!='-')))

        factory_items=[{**r} for r in results if (r['stF'] or 0)>0]
        bangwa_items=[{**r} for r in results if (r['stF'] or 0)==0]

        col1,col2=st.columns(2)
        with col1:
            st.metric("🏭 โรงงาน", f"{sum(i['qty'] for i in factory_items):,} คู่")
        with col2:
            st.metric("🏬 บางหว้า", f"{sum(i['qty'] for i in bangwa_items):,} คู่")

        st.divider()
        c1,c2=st.columns(2)
        with c1:
            if factory_items:
                xlsx_f=build_excel(factory_items,True,all_so,all_po,all_dates)
                st.download_button("⬇ ดาวน์โหลด Excel โรงงาน",xlsx_f,
                    f"barcode_factory_{datetime.date.today()}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.session_state['factory_xlsx']=xlsx_f
                st.session_state['factory_items']=factory_items
        with c2:
            if bangwa_items:
                xlsx_b=build_excel(bangwa_items,False,all_so,all_po,all_dates)
                st.download_button("⬇ ดาวน์โหลด Excel บางหว้า",xlsx_b,
                    f"barcode_bangwa_{datetime.date.today()}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.session_state['bangwa_xlsx']=xlsx_b
                st.session_state['bangwa_items']=bangwa_items

with tab3:
    if 'results' not in st.session_state:
        st.info("กรุณาตรวจสอบสต็อคก่อนใน Tab แรกครับ")
    else:
        st.subheader("⚙️ ตั้งค่าอีเมล")
        with st.form("email_form"):
            c1,c2=st.columns(2)
            with c1:
                from_email=st.text_input("อีเมลผู้ส่ง (Gmail)", placeholder="your@gmail.com")
                app_password=st.text_input("App Password", type="password", placeholder="xxxx xxxx xxxx xxxx")
            with c2:
                to_factory=st.text_input("ถึงโรงงาน (To)", placeholder="factory@example.com")
                to_bangwa=st.text_input("ถึงบางหว้า (To)", placeholder="warehouse@example.com")
            cc=st.text_input("CC (ถ้ามี)")
            submitted=st.form_submit_button("📧 ส่งอีเมลทันที", type="primary", use_container_width=True)

        if submitted and from_email and app_password:
            results=st.session_state['results']
            all_so=', '.join(sorted(set(r['soNo'] for r in results)))
            all_dates=', '.join(sorted(set(r['delivery'] for r in results if r['delivery']!='-')))
            today=datetime.date.today().strftime('%d/%m/%Y')

            def send(to_addr, subject, body_txt, xlsx_bytes, fname):
                msg=MIMEMultipart()
                msg['From']=from_email; msg['To']=to_addr; msg['Subject']=subject
                if cc: msg['Cc']=cc
                msg.attach(MIMEText(body_txt,'plain','utf-8'))
                part=MIMEBase('application','octet-stream'); part.set_payload(xlsx_bytes)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',f'attachment; filename="{fname}"')
                msg.attach(part)
                ctx=ssl.create_default_context()
                pw=app_password.replace(' ','')
                with smtplib.SMTP('smtp.gmail.com',587) as srv:
                    srv.ehlo(); srv.starttls(context=ctx); srv.ehlo()
                    srv.login(from_email,pw)
                    recip=[to_addr]+([cc] if cc else [])
                    srv.sendmail(from_email,recip,msg.as_string())

            sent=[]
            if to_factory and 'factory_xlsx' in st.session_state:
                with st.spinner("กำลังส่งอีเมลโรงงาน..."):
                    try:
                        send(to_factory,
                             f'รายการสินค้าติดบาร์โค้ด {all_so} — ส่ง {all_dates}',
                             f'เรียน ทางโรงงาน\n\nขอแจ้งรายการสินค้าสำหรับติดบาร์โค้ด\nSO: {all_so}\nกำหนดส่ง: {all_dates}\n\nขอแสดงความนับถือ\nฝ่าย Sale Support — นันยางมาร์เก็ตติ้ง จำกัด',
                             st.session_state['factory_xlsx'],
                             f'barcode_factory_{datetime.date.today()}.xlsx')
                        sent.append('✅ ส่งอีเมลโรงงานสำเร็จ')
                    except Exception as e:
                        sent.append(f'❌ ส่งโรงงานไม่ได้: {e}')

            if to_bangwa and 'bangwa_xlsx' in st.session_state:
                with st.spinner("กำลังส่งอีเมลบางหว้า..."):
                    try:
                        send(to_bangwa,
                             f'รายการสินค้าหยิบออกจากโกดัง {all_so} — ส่ง {all_dates}',
                             f'เรียน ทีมโกดังบางหว้า\n\nขอแจ้งรายการสินค้าที่ต้องหยิบออกจากโกดัง\nSO: {all_so}\nกำหนดส่ง: {all_dates}\n\nขอแสดงความนับถือ\nฝ่าย Sale Support — นันยางมาร์เก็ตติ้ง จำกัด',
                             st.session_state['bangwa_xlsx'],
                             f'barcode_bangwa_{datetime.date.today()}.xlsx')
                        sent.append('✅ ส่งอีเมลบางหว้าสำเร็จ')
                    except Exception as e:
                        sent.append(f'❌ ส่งบางหว้าไม่ได้: {e}')

            for s in sent:
                if '✅' in s: st.success(s)
                else: st.error(s)
